import requests
from datetime import datetime, date
import time
import concurrent.futures
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from openpyxl import load_workbook


def is_date_in_range(gmt_commit: str, start_date: str = None, end_date: str = None) -> bool:
    """
    检查提交时间是否在指定的时间范围内

    Args:
        gmt_commit: 提交时间字符串，格式为 "YYYY-MM-DD HH:mm:ss"
        start_date: 开始日期，格式为 "YYYY-MM-DD"，如果为 None 则不限制开始时间
        end_date: 结束日期，格式为 "YYYY-MM-DD"，如果为 None 则不限制结束时间

    Returns:
        bool: 如果在时间范围内返回 True，否则返回 False
    """
    if not gmt_commit:
        return False

    try:
        # 提取日期部分
        commit_date = gmt_commit.split()[0] if ' ' in gmt_commit else gmt_commit
        commit_datetime = datetime.strptime(commit_date, "%Y-%m-%d")

        # 检查开始日期
        if start_date:
            start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
            if commit_datetime < start_datetime:
                return False

        # 检查结束日期
        if end_date:
            end_datetime = datetime.strptime(end_date, "%Y-%m-%d")
            if commit_datetime > end_datetime:
                return False

        return True
    except (ValueError, AttributeError) as e:
        print(f"日期格式错误：{e}")
        return False

def retry_on_failure(max_retries=3, delay=1):
    """
    重试装饰器，用于处理网络请求失败的情况
    
    Args:
        max_retries: 最大重试次数
        delay: 重试间隔（秒）
    """
    def decorator(func):
        def wrapper(*args, **kwargs):
            retries = 0
            while retries < max_retries:
                try:
                    return func(*args, **kwargs)
                except requests.exceptions.RequestException as e:
                    retries += 1
                    if retries >= max_retries:
                        raise
                    print(f"请求失败，{delay}秒后重试 ({retries}/{max_retries})...")
                    print(f"错误信息: {e}")
                    time.sleep(delay)
        return wrapper
    return decorator

# 常量定义
API_BASE_URL = "https://labelx.alibaba-inc.com"
SUBTASKS_API = f"{API_BASE_URL}/api/v1/label/center/subTasks"
SUBTASK_DETAIL_API = f"{API_BASE_URL}/api/v1/label/center/subTask/{{subtask_id}}/data"
LOGIN_URL = f"{API_BASE_URL}/corpora/labeling/labelingTask?projectId=1023"

# 默认请求参数
DEFAULT_PARAMS = {
    "type": "label",
    "keyword": "",
    "appId": 1023,
    "finished": "true",
    "page": 1,
    "pageSize": 50
}

# 默认详情请求参数
DEFAULT_DETAIL_PARAMS = {
    "page": 1,
    "pageSize": 50,
    "filterPassedVote": "false",
    "filter": '{"questions":[],"dataStatus":"ALL","questionsQueryConditions":"AND"}'
}

# 默认请求头
DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "X-Requested-With": "XMLHttpRequest"
}


def build_headers(cookie_dict: dict) -> dict:
    """
    构建请求头
    
    Args:
        cookie_dict: Cookie 字典
    
    Returns:
        完整的请求头字典
    """
    headers = DEFAULT_HEADERS.copy()
    headers["Cookie"] = "; ".join([f"{k}={v}" for k, v in cookie_dict.items()])
    return headers


def build_params(base_params: dict) -> dict:
    """
    构建请求参数，添加时间戳
    
    Args:
        base_params: 基础参数
    
    Returns:
        包含时间戳的完整参数
    """
    params = base_params.copy()
    params["_"] = int(time.time() * 1000)
    return params


def get_cookies_from_browser() -> tuple:
    """
    使用 Selenium 通过调试端口或自动启动方式获取 Edge 浏览器 Cookie 和用户名

    Returns:
        tuple: (cookie_dict, username)，如果获取失败返回 ({}, None)
    """
    print("=" * 60)
    print("获取 Cookie 的两种方式：")
    print("1. 连接已打开的 Edge 浏览器（调试模式 127.0.0.1:9222）")
    print("2. 自动启动新的 Edge 浏览器")
    print("=" * 60)

    # 方式 1：尝试连接已打开的 Edge 浏览器（调试模式）
    try:
        print("\n[方式 1] 尝试连接 Edge (调试模式 127.0.0.1:9222)...")
        options = Options()
        options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        driver = webdriver.Edge(options=options)

        print("✓ 成功连接到 Edge 浏览器")
        driver.get(LOGIN_URL)

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            print("页面已加载")
        except TimeoutException:
            print("页面加载超时，继续执行")

        print("正在获取 Cookie...")
        cookies = driver.get_cookies()
        cookie_dict = {cookie['name']: cookie['value'] for cookie in cookies}
        print(f"成功获取 {len(cookies)} 个 Cookie")
        driver.quit()

        return cookie_dict, None

    except WebDriverException as e:
        print(f"✗ 连接调试模式失败：{e}")
        print("\n[方式 2] 自动启动新的 Edge 浏览器...")

        edge_options = Options()
        edge_options.add_argument("--disable-gpu")
        edge_options.add_argument("--no-sandbox")
        edge_options.add_argument("--disable-dev-shm-usage")
        edge_options.add_argument("--window-size=1920,1080")

        try:
            print("正在初始化 EdgeDriver...")
            driver = webdriver.Edge(service=Service(), options=edge_options)

            print("Edge 浏览器已启动，正在访问页面...")
            driver.get(LOGIN_URL)

            wait = WebDriverWait(driver, 15)
            try:
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                print("页面已加载")
            except TimeoutException:
                print("等待页面超时，继续执行...")

            if "login" in driver.current_url.lower():
                print("警告：检测到未登录状态，请先在浏览器中登录！")
                print("请在打开的浏览器中完成登录，然后按回车继续...")
                input()
                try:
                    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    print("登录成功，页面已加载")
                except TimeoutException:
                    print("登录后页面加载超时，继续执行...")

            print("正在获取 Cookie...")
            cookies = driver.get_cookies()
            cookie_dict = {cookie['name']: cookie['value'] for cookie in cookies}
            print(f"成功获取 {len(cookies)} 个 Cookie")
            driver.quit()

            return cookie_dict, None

        except Exception as e:
            print(f"获取 Cookie 失败：{e}")
            return {}, None
        finally:
            try:
                driver.quit()
            except:
                pass

@retry_on_failure(max_retries=3, delay=1)
def get_subtask_detail(cookie_dict: dict, subtask_id: int) -> dict:
    """
    获取子任务的详细信息

    通过调用详细数据 API 获取单个子任务的完整信息，包括 dataList 等详细内容

    Args:
        cookie_dict: Cookie 字典，包含认证信息
        subtask_id: 子任务 ID

    Returns:
        子任务的详细信息字典，包含 id、taskId、batchId、taskName、dataList 等字段
        如果获取失败则返回 None
    """
    detail_url = SUBTASK_DETAIL_API.format(subtask_id=subtask_id)
    detail_params = build_params(DEFAULT_DETAIL_PARAMS)
    headers = build_headers(cookie_dict)

    try:
        response = requests.get(detail_url, params=detail_params, headers=headers)
        response.raise_for_status()

        json_data = response.json()

        if json_data.get("code") == 0 and json_data.get("success"):
            return json_data.get("data", {})
        else:
            print(f"获取子任务 {subtask_id} 详情失败：{json_data.get('message')}")
            return None

    except Exception as e:
        print(f"获取子任务详情失败：{e}")
        return None


def fetch_all_pages(cookie_dict: dict, target_date: str = None, max_workers: int = 5, start_date: str = None, end_date: str = None) -> list:
    """
    分页获取所有符合条件的数据

    遍历所有页面，筛选符合条件的子任务（rejectReason 为 null 且 gmtCommit 匹配指定日期或时间范围），
    并获取每个子任务的详细信息，包括 dataList 中的 wav_id、duration、value 等

    Args:
        cookie_dict: Cookie 字典，包含认证信息
        target_date: 目标日期，格式为 "YYYY-MM-DD"
                    如果为 None 则不限制日期，抓取所有数据
        max_workers: 并发工作线程数
        start_date: 开始日期，格式为 "YYYY-MM-DD"，如果为None则不限制开始时间
        end_date: 结束日期，格式为 "YYYY-MM-DD"，如果为None则不限制结束时间

    Returns:
        包含所有符合条件子任务的列表，每个子任务包含：
        - id: 子任务 ID
        - taskId: 任务 ID
        - batchId: 批次 ID
        - taskName: 任务名称
        - gmtCommit: 提交时间
        - dataList: 数据列表，包含 wav_id、duration、value 等
    """
    all_results = []
    page = 1
    headers = build_headers(cookie_dict)

    # 分页循环获取数据
    while True:
        current_params = DEFAULT_PARAMS.copy()
        current_params["page"] = page
        current_params = build_params(current_params)

        try:
            print(f"请求第 {page} 页...")
            # 添加重试逻辑
            retries = 0
            max_retries = 3
            while retries < max_retries:
                try:
                    response = requests.get(SUBTASKS_API, params=current_params, headers=headers)
                    response.raise_for_status()
                    break
                except requests.exceptions.RequestException as e:
                    retries += 1
                    if retries >= max_retries:
                        raise
                    print(f"请求失败，1秒后重试 ({retries}/{max_retries})...")
                    print(f"错误信息: {e}")
                    time.sleep(1)

            json_data = response.json()

            if json_data.get("code") != 0 or not json_data.get("success"):
                print(f"请求失败：{json_data.get('message')}")
                break

            data_list = json_data.get("data", {}).get("data", [])
            total_count = json_data.get("data", {}).get("recordCount", 0)

            if page == 1:
                print(f"总记录数：{total_count}")

            if not data_list:
                break

            print(f"正在处理第 {page} 页，共 {len(data_list)} 条数据...")

            # 收集符合条件的子任务 ID
            subtask_ids = []
            for item in data_list:
                # 过滤掉有拒绝原因的数据
                if item.get("rejectReason") is not None:
                    continue

                # 获取提交时间
                gmt_commit = item.get("gmtCommit")
                if not gmt_commit:
                    continue

                # 检查时间条件
                # 优先使用时间范围筛选，如果没有时间范围则使用单个日期筛选
                if start_date or end_date:
                    # 使用时间范围筛选
                    if not is_date_in_range(gmt_commit, start_date, end_date):
                        continue
                elif target_date:
                    # 使用单个日期筛选
                    if not gmt_commit.startswith(target_date):
                        continue

                # 获取子任务的 ID
                subtask_id = item.get("id")
                subtask_ids.append(subtask_id)

                        # 并发获取子任务详情
            if subtask_ids:
                print(f"  → 并发获取 {len(subtask_ids)} 个子任务详情...")
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                    future_to_subtask = {
                        executor.submit(get_subtask_detail, cookie_dict, subtask_id): subtask_id
                        for subtask_id in subtask_ids
                    }

                    for future in concurrent.futures.as_completed(future_to_subtask):
                        subtask_id = future_to_subtask[future]
                        try:
                            detail_data = future.result()
                            if detail_data:
                                filtered_item = {
                                    "id": detail_data.get("id"),
                                    "taskId": detail_data.get("taskId"),
                                    "batchId": detail_data.get("batchId"),
                                    "taskName": detail_data.get("taskName"),
                                    "gmtCommit": detail_data.get("gmtCommit"),
                                    "dataList": []
                                }

                                for data_item in detail_data.get("dataList", []):
                                    data_data = data_item.get("data", {})
                                    result_data = data_item.get("result", {})
                                    mark_result = result_data.get("markResult", [])
                                    filtered_item["dataList"].append({
                                        "wav_id": data_data.get("wav_id"),
                                        "duration": data_data.get("duration"),
                                        "value": mark_result[0].get("value", []) if mark_result else []
                                    })

                                all_results.append(filtered_item)
                                print(f"  ✓ 已获取：{detail_data.get('taskName')} (dataList: {len(detail_data.get('dataList', []))} 条)")
                        except Exception as e:
                            print(f"  ✗ 获取子任务 {subtask_id} 详情失败：{e}")

            # 判断是否还有下一页
            if len(data_list) < current_params["pageSize"]:
                break

            page += 1

        except requests.exceptions.RequestException as e:
            print(f"请求错误：{e}")
            break
        except Exception as e:
            print(f"处理错误：{e}")
            break

    return all_results


def main():
    """主函数"""
    default_max_workers = 20  # 增加并发数
    page_size = 50

    print("\n请输入要查询的时间范围：")
    print("1. 单个日期（YYYY-MM-DD），如：2026-03-26")
    print("2. 时间范围（YYYY-MM-DD YYYY-MM-DD），如：2026-03-01 2026-03-31")
    print("3. 回车查询所有日期")
    user_input = input("请输入：").strip()

    target_date = start_date = end_date = ""

    if user_input:
        parts = user_input.split()
        if len(parts) == 1:
            target_date = parts[0]
        elif len(parts) == 2:
            start_date, end_date = parts[0], parts[1]
        else:
            print("输入格式错误！")
            return

    if target_date:
        try:
            datetime.strptime(target_date, "%Y-%m-%d")
            print(f"将筛选 {target_date} 提交的数据")
        except ValueError:
            print("日期格式错误！")
            return

    if start_date and end_date:
        try:
            if datetime.strptime(start_date, "%Y-%m-%d") > datetime.strptime(end_date, "%Y-%m-%d"):
                print("开始日期不能大于结束日期")
                return
            print(f"将筛选 {start_date} 到 {end_date} 之间提交的数据")
        except ValueError:
            print("日期格式错误！")
            return

    # 让用户手动输入账号名
    username = input("\n请输入您的账号名（用于文件名标识）：").strip()
    if not username:
        print("未输入账号名，将使用默认文件名")

    DEFAULT_PARAMS["pageSize"] = page_size
    DEFAULT_DETAIL_PARAMS["pageSize"] = page_size

    print(f"\n开始抓取数据 - {date.today()}")
    print("=" * 60)

    start_time = time.time()
    cookie_dict, _ = get_cookies_from_browser()

    if not cookie_dict:
        return

    results = fetch_all_pages(cookie_dict, target_date, default_max_workers, start_date, end_date)

    if not results:
        print("未找到符合条件的数据")
        return

    total_time = time.time() - start_time

    print(f"\n找到 {len(results)} 条符合条件的数据：")
    print("=" * 60)
    
    # 统计总数据条数
    total_records = sum(len(item.get('dataList', [])) for item in results)
    print(f"共 {total_records} 条记录")

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    date_suffix = f"{start_date}_to_{end_date}" if start_date and end_date else (target_date if target_date else "all")
    
    # 使用用户输入的账号名生成 Excel 文件名
    if username:
        # 移除文件名中的非法字符：/ \ : * ? " < > |
        safe_username = "".join(c for c in username if c not in r'<>:"/\|？*')
        excel_filename = f"ASR 数据_{safe_username}_{date_suffix}_{timestamp}.xlsx"
    else:
        excel_filename = f"ASR 数据_{date_suffix}_{timestamp}.xlsx"

    # 读取 Excel 模板并填入数据
    try:
        import os
        template_path = '新表.xlsx'
        print(f"正在查找模板文件：{os.path.abspath(template_path)}")
        
        if not os.path.exists(template_path):
            print(f"警告：未找到模板文件 '{template_path}'，请确保它与程序在同一目录下")
            print("跳过 Excel 导出")
        else:
            print("找到模板文件，开始写入数据...")
            wb = load_workbook(template_path)
            ws = wb.active
        
        # 准备数据
        excel_date = target_date if target_date else (start_date if start_date else date.today().isoformat())
        row_data = []
        
        for item in results:
            for data_item in item.get('dataList', []):
                # B:日期，D:账号名，E:任务名称，F:任务 ID, G:子任务 ID, H:分包 ID, I:内容 id, K:音频时长，L:是否有效
                # 将 duration 转换为数字
                duration_val = data_item.get('duration', 0)
                if duration_val:
                    try:
                        duration_val = float(duration_val)
                    except (ValueError, TypeError):
                        duration_val = 0
                
                # 从 data_item 中直接获取 value（是否有效）
                is_valid = data_item.get('value', '')
                if isinstance(is_valid, list) and is_valid:
                    is_valid = is_valid[0]
                
                row = [
                    excel_date,  # B 列 - 日期
                    username or '',  # D 列 - 账号名
                    item.get('taskName', ''),  # E 列 - 任务名称
                    item.get('taskId', ''),  # F 列 - 任务 ID
                    item.get('id', ''),  # G 列 - 子任务 ID
                    item.get('batchId', ''),  # H 列 - 分包 ID
                    data_item.get('wav_id', ''),  # I 列 - 内容 id
                    duration_val,  # K 列 - 音频时长
                    is_valid,  # L 列 - 是否有效
                ]
                row_data.append(row)
        
        # 使用切片赋值一次性填入数据
        if row_data:
            # 从第 5 行开始填入（第 5 行是第一条数据行）
            for i, row in enumerate(row_data):
                start_row = 5 + i
                ws[f'B{start_row}'] = row[0]  # 日期
                ws[f'D{start_row}'] = row[1]  # 账号名
                ws[f'E{start_row}'] = row[2]  # 任务名称
                ws[f'F{start_row}'] = row[3]  # 任务 ID
                ws[f'G{start_row}'] = row[4]  # 子任务 ID
                ws[f'H{start_row}'] = row[5]  # 分包 ID
                ws[f'I{start_row}'] = row[6]  # 内容 id
                ws[f'K{start_row}'] = row[7]  # 音频时长
                ws[f'L{start_row}'] = row[8]  # 是否有效
            
            # 更新统计表（第 3 行）
            ws['A3'] = len(row_data)  # 条数
            
            # 计算总时长（只统计音频时长）
            total_duration = sum(float(row[7]) for row in row_data if row[7])
            ws['K3'] = total_duration / 60 if total_duration else 0  # 当日音频时长 (m)
        
        # 保存 Excel 文件
        wb.save(excel_filename)
        print(f"Excel 数据已保存到：{os.path.abspath(excel_filename)}")
        
    except Exception as e:
        import traceback
        print(f"写入 Excel 时出错：{e}")
        traceback.print_exc()
    print("=" * 60)
    print(f"抓取完成，共 {len(results)} 条数据")
    print(f"\n抓取速度统计：")
    print(f"  总耗时：{total_time:.2f} 秒")
    if total_time > 0:
        print(f"  平均速度：{len(results) / total_time:.2f} 条/秒")
    print("=" * 60)

if __name__ == "__main__":
    main()
